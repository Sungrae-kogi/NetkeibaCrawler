import imaplib
import email
import json
import os
import smtplib
import pandas as pd
import pymysql
import re
from email.header import decode_header
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from pathlib import Path

# 이 파일 기준으로 프로젝트 루트 경로 계산
# src/reporting/Reporting/ -> src/reporting/ -> src/ -> project_root/
_BASE_DIR = Path(__file__).resolve().parent
_PROJECT_ROOT = _BASE_DIR.parent.parent.parent
from io import StringIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

def get_credentials():
    config_path = _PROJECT_ROOT / "config" / "naverworks_config.json"
    if not config_path.exists():
        return None, None
    with open(config_path, "r", encoding="utf-8") as f:
        config = json.load(f)
        return config.get("email"), config.get("password")

def get_db_connection():
    import os
    env = os.environ.get("APP_ENV", "prod")
    config_name = "db_config_test.json" if env == "test" else "db_config.json"
    config_path = _PROJECT_ROOT / "config" / config_name
    with open(config_path, "r", encoding="utf-8") as f:
        config = json.load(f)
    return pymysql.connect(
        host=config['host'],
        port=config.get('port', 3306),
        user=config['user'],
        password=config['password'],
        database="k_mafeel",
        cursorclass=pymysql.cursors.DictCursor
    )

def decode_mime_words(s):
    if not s:
        return ""
    decoded_words = decode_header(s)
    result = []
    for word, charset in decoded_words:
        if isinstance(word, bytes):
            result.append(word.decode(charset or 'utf-8', errors='replace'))
        else:
            result.append(word)
    return "".join(result)

def fetch_prediction_table(venue_kor, date_str=None):
    """
    naver_works 메일함에서 mafeel@becurio.com이 보낸 [서울-날짜]경주결과 Report 제목의 메일을 읽어옵니다.
    날짜(date_str)가 지정되지 않은 경우 가장 최신 메일에서 날짜를 자동 추출합니다.
    """
    user_email, user_password = get_credentials()
    imap_server = "imap.worksmobile.com"
    
    mail = imaplib.IMAP4_SSL(imap_server, 993)
    mail.login(user_email, user_password)
    mail.select("mafeel")
    
    status, messages = mail.search(None, 'FROM', '"mafeel@becurio.com"')
    mail_ids = messages[0].split()
    
    # [서울-20260516]경주결과 Report 형태 검색 필터 생성
    if date_str:
        target_subject_part = f"[{venue_kor}-{date_str}]경주결과 Report"
    else:
        target_subject_part = f"[{venue_kor}-"
    
    df = None
    found_date = date_str
    
    for i in reversed(mail_ids):
        res, msg_data = mail.fetch(i, "(RFC822)")
        for response_part in msg_data:
            if isinstance(response_part, tuple):
                msg = email.message_from_bytes(response_part[1])
                subject = decode_mime_words(msg["Subject"])
                
                # 메일 제목에 target_subject_part가 포함되고 "경주결과 Report"가 들어있는지 검사
                if subject and target_subject_part in subject and "경주결과 Report" in subject:
                    if not date_str:
                        # 메일 제목에서 날짜(8자리 숫자) 자동 추출
                        m = re.search(rf"\[{venue_kor}-(\d{{8}})\]", subject)
                        if m:
                            found_date = m.group(1)
                        else:
                            continue # 날짜 패턴이 매칭되지 않으면 패스
                            
                    for part in msg.walk():
                        if part.get_content_maintype() == 'multipart': continue
                        if part.get_content_type() == "text/html" and part.get('Content-Disposition') is None:
                            try:
                                html_body = part.get_payload(decode=True).decode('utf-8', errors='replace')
                                tables = pd.read_html(StringIO(html_body), header=0)
                                for table in tables:
                                    # 예측 데이터 테이블 컬럼 확인 (순위, 마명 등)
                                    if '순위' in table.columns or '한글마명' in table.columns or '마명' in table.columns:
                                        df = table
                                        break
                                if df is not None:
                                    break
                            except Exception:
                                pass
                    if df is not None:
                        mail.logout()
                        return df, found_date
    mail.logout()
    return None, found_date

def generate_hybrid_report(pred_df, venue_kor, date_str):
    """
    예측 데이터와 api_race_result 테이블의 실제 경기 결과를 병합하여 리포트 데이터를 생성합니다.
    """
    # 컬럼 표준화
    if '마명' in pred_df.columns and '한글마명' not in pred_df.columns:
        pred_df = pred_df.rename(columns={'마명': '한글마명'})
        
    keep_cols = ['경주정보', '날짜', '경주번호', '순위', '출발번호', '한글마명']
    actual_keep = [col for col in keep_cols if col in pred_df.columns]
    df = pred_df[actual_keep].copy()
    
    df.rename(columns={'순위': '[예측]순위', '출발번호': '[예측]출발번호', '한글마명': '[예측]한글마명'}, inplace=True)
    
    # date_str을 int 타입으로 안전 변환 (예: '2026-05-10' 또는 '20260510' -> 20260510)
    try:
        date_int = int(str(date_str).replace('-', ''))
    except Exception:
        date_int = 0
        
    conn = get_db_connection()
    try:
        actual_data = []
        pred_odds_list = []
        pred_actual_ranks = []
        
        for _, row in df.iterrows():
            race_no = row.get('경주번호', 1)
            pred_rank = row.get('[예측]순위', '')
            pred_horse_no = row.get('[예측]출발번호', '')
            
            # 경주번호 포맷 정제 (예: '3경주' 또는 3 -> 3)
            race_str = str(race_no).replace('경주', '').strip()
            try:
                race_num = int(re.sub(r'\D', '', race_str))
            except Exception:
                race_num = 1
                
            # 출발번호(마번) 정수화 (예: '03' -> 3)
            try:
                pred_horse_int = int(pred_horse_no)
            except Exception:
                pred_horse_int = 0
            
            with conn.cursor() as cursor:
                # 1. 예측마의 배당률 및 실제 순위 가져오기 (RACE_DT, RACE_NO, GTNO 모두 int 타입으로 매칭)
                sql_pred = "SELECT WIN_PRICE, RK FROM api_race_result WHERE RACE_DT=%s AND RCCRS_NM=%s AND RACE_NO=%s AND GTNO=%s"
                cursor.execute(sql_pred, (date_int, venue_kor, race_num, pred_horse_int))
                res_pred = cursor.fetchone()
                if res_pred:
                    pred_odds_list.append(res_pred['WIN_PRICE'])
                    pred_actual_rank = f"{res_pred['RK']}등" if res_pred['RK'] else '-'
                else:
                    pred_odds_list.append('-')
                    pred_actual_rank = '-'
                pred_actual_ranks.append(pred_actual_rank)

            with conn.cursor() as cursor:
                # 2. 실제 순위 정보 대조 (1등, 2등, 3등 포디움 채우기)
                if pred_rank in ['1등', '2등', '3등']:
                    try:
                        rank_num = int(pred_rank.replace('등', ''))
                    except Exception:
                        rank_num = 1
                        
                    sql_actual = "SELECT RK, GTNO, HRNM, WIN_PRICE FROM api_race_result WHERE RACE_DT=%s AND RCCRS_NM=%s AND RACE_NO=%s AND RK=%s"
                    cursor.execute(sql_actual, (date_int, venue_kor, race_num, rank_num))
                    res_act = cursor.fetchone()
                    if res_act:
                        actual_data.append({
                            '[실제]순위': f"{res_act['RK']}등", 
                            '[실제]출발번호': res_act['GTNO'], 
                            '[실제]한글마명': res_act['HRNM'], 
                            '[실제]배당률': res_act['WIN_PRICE']
                        })
                    else:
                        actual_data.append({'[실제]순위': '-', '[실제]출발번호': '-', '[실제]한글마명': '-', '[실제]배당률': '-'})
                else:
                    # 복병마 행은 포디움 칸 비워둠
                    actual_data.append({'[실제]순위': '-', '[실제]출발번호': '-', '[실제]한글마명': '-', '[실제]배당률': '-'})
    finally:
        conn.close()
        
    df['[예측]배당률'] = pred_odds_list
    df['[예측]실제순위'] = pred_actual_ranks
    actual_df = pd.DataFrame(actual_data)
    final_df = pd.concat([df.reset_index(drop=True), actual_df.reset_index(drop=True)], axis=1)
    return final_df

def create_excel_file(df, filepath):
    wb = Workbook()
    ws = wb.active
    ws.title = "결과보고서"
    
    headers_row1 = ["경주정보", "날짜", "경주번호", 
                    "예측 (Prediction)", "", "", "", "",
                    "실제 결과 (Actual)", "", "", ""]
    headers_row2 = ["", "", "", 
                    "순위", "출발번호", "한글마명", "배당률", "실제순위",
                    "순위", "출발번호", "한글마명", "배당률"]
    ws.append(headers_row1)
    ws.append(headers_row2)
    
    for _, row in df.iterrows():
        ws.append([
            row.get("경주정보", ""), row.get("날짜", ""), row.get("경주번호", ""),
            row.get("[예측]순위", ""), row.get("[예측]출발번호", ""), row.get("[예측]한글마명", ""), row.get("[예측]배당률", ""), row.get("[예측]실제순위", ""),
            row.get("[실제]순위", ""), row.get("[실제]출발번호", ""), row.get("[실제]한글마명", ""), row.get("[실제]배당률", "")
        ])
        
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_header = PatternFill("solid", fgColor="EFEFEF")
    fill_hit = PatternFill("solid", fgColor="FFFF00") # 노란색 강조
    align_center = Alignment(horizontal="center", vertical="center")
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=12):
        for cell in row:
            cell.border = border
            cell.alignment = align_center
            if cell.row in [1, 2]:
                cell.fill = fill_header
                cell.font = Font(bold=True)
                
    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:B2")
    ws.merge_cells("C1:C2")
    ws.merge_cells("D1:H1") # 예측 영역 머지 (5개 열)
    ws.merge_cells("I1:L1") # 실제 결과 영역 머지 (4개 열)
    
    thick_bottom = Side(border_style="medium", color="000000")
                
    start_row = 3
    for i in range(3, ws.max_row + 2):
        if i > ws.max_row or ws.cell(row=i, column=3).value != ws.cell(row=i-1, column=3).value:
            end_row = i - 1
            
            # 예측마가 실제 1, 2, 3등 안에 들었으면 해당 행 노란색 강조
            for r in range(start_row, end_row + 1):
                actual_rank = str(ws.cell(row=r, column=8).value) # Column H: [예측]실제순위
                if actual_rank in ['1등', '2등', '3등']:
                    # 예측 컬럼들 D~H(4~8열) 노란색 칠하기
                    for col in range(4, 9):
                        ws.cell(row=r, column=col).fill = fill_hit

            # 경주 정보 셀 병합 (경주번호 기준)
            if end_row > start_row:
                ws.merge_cells(start_row=start_row, end_row=end_row, start_column=1, end_column=1)
                ws.merge_cells(start_row=start_row, end_row=end_row, start_column=2, end_column=2)
                ws.merge_cells(start_row=start_row, end_row=end_row, start_column=3, end_column=3)
                
            # 경주별 마지막 행(end_row)에 굵은 아래쪽 테두리 추가하여 시각적 구분 강화
            for col in range(1, 13):
                current_border = ws.cell(row=end_row, column=col).border
                ws.cell(row=end_row, column=col).border = Border(
                    left=current_border.left,
                    right=current_border.right,
                    top=current_border.top,
                    bottom=thick_bottom
                )
                
            start_row = i
            
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15 # 실제순위
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 12
            
    wb.save(filepath)

def send_report_email(final_df, to_emails, venue_kor, date_str, cc_emails=None):
    user_email, user_password = get_credentials()
    smtp_server = "smtp.worksmobile.com"
    
    msg = MIMEMultipart()
    msg['From'] = user_email
    
    if isinstance(to_emails, list):
        msg['To'] = ", ".join(to_emails)
    else:
        msg['To'] = to_emails
        
    if cc_emails:
        if isinstance(cc_emails, list):
            msg['Cc'] = ", ".join(cc_emails)
        else:
            msg['Cc'] = cc_emails
        
    msg['Subject'] = f"[{venue_kor}-{date_str}] 예측과 실제 결과 비교 리포트"
    
    # HTML 테이블 생성 (노란색 하이라이트 반영)
    html_lines = ['<table border="1" style="border-collapse: collapse; width: 100%; text-align: center;">', "  <thead>", "    <tr>"]
    for col in final_df.columns:
        html_lines.append(f'      <th style="background-color: #f2f2f2; padding: 8px;">{col}</th>')
    html_lines.append("    </tr>")
    html_lines.append("  </thead>")
    html_lines.append("  <tbody>")
    
    total_rows = len(final_df)
    for idx, (_, row) in enumerate(final_df.iterrows()):
        html_lines.append("    <tr>")
        actual_rank = str(row.get('[예측]실제순위', ''))
        is_hit = actual_rank in ['1등', '2등', '3등']
        
        # 다음 행의 경주번호가 다른지 판단 (마지막 행이거나 다음 경주번호와 다르면 경주 종료선 표시)
        is_last_of_race = False
        if idx == total_rows - 1:
            is_last_of_race = True
        else:
            next_race = final_df.iloc[idx + 1].get('경주번호')
            current_race = row.get('경주번호')
            if next_race != current_race:
                is_last_of_race = True
        
        for col in final_df.columns:
            val = row[col] if pd.notna(row[col]) else ''
            
            # 스타일 요소 정의
            styles = ["padding: 8px;"]
            if is_hit and col in ['[예측]순위', '[예측]출발번호', '[예측]한글마명', '[예측]배당률', '[예측]실제순위']:
                styles.append("background-color: #FFFF00;")
                
            if is_last_of_race:
                styles.append("border-bottom: 3px solid #000000;")
                
            style_str = f' style="{" ".join(styles)}"'
            html_lines.append(f"      <td{style_str}>{val}</td>")
        html_lines.append("    </tr>")
    html_lines.append("  </tbody>")
    html_lines.append("</table>")
    table_html = "\n".join(html_lines)
    
    date_hyphen = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
    html_content = f"""
    <html>
    <head>
        <style>
            table {{
                border-collapse: collapse;
                width: 100%;
                max-width: 1000px;
                font-family: 'Malgun Gothic', sans-serif;
                font-size: 12px;
                border: 1px solid #dddddd;
                margin-top: 15px;
                margin-bottom: 15px;
            }}
            th, td {{
                padding: 8px;
                text-align: center;
                border: 1px solid #dddddd;
            }}
            th {{
                background-color: #f2f2f2;
                font-weight: bold;
                color: #333333;
            }}
            tr:nth-child(even) {{
                background-color: #f9f9f9;
            }}
        </style>
    </head>
    <body>
        <p>{date_hyphen} {venue_kor} 경기 결과와 예측 비교 리포트를 첨부 파일로 송부합니다.</p>
        <br/>
        <h3>📊 [결과 비교 요약 표]</h3>
        {table_html}
        <br/>
        <p>감사합니다.</p>
    </body>
    </html>
    """
    msg.attach(MIMEText(html_content, 'html'))
    
    excel_filename = f"{venue_kor}_{date_str}_예측과_실제_결과_비교_리포트.xlsx"
    create_excel_file(final_df, excel_filename)
    
    with open(excel_filename, "rb") as f:
        attach = MIMEApplication(f.read(), _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        attach.add_header('Content-Disposition', 'attachment', filename=excel_filename)
        msg.attach(attach)
    
    try:
        server = smtplib.SMTP_SSL(smtp_server, 465)
        server.login(user_email, user_password)
        server.send_message(msg)
        server.quit()
        print(f"[{venue_kor}-{date_str}] 이메일 발송 성공!")
    except Exception as e:
        print(f"이메일 발송 실패: {e}")

def run_kor_reporting_pipeline(venue_kor, date_str=None, to_emails=None, cc_emails=None):
    """
    한국 전용 결과 리포트 파이프라인 메인 실행부
    """
    import os
    env = os.environ.get("APP_ENV", "prod")
    if env == "test":
        to_emails = ["whvkek@becurio.com"]
        cc_emails = None
        print("[TEST MODE] 한국 리포트 이메일 수신자 격리 활성화 -> whvkek@becurio.com 수신 전용")
    else:
        if to_emails is None:
            to_emails = "ysoh@becurio.com"
        
        if cc_emails is None:
            cc_emails = ["pizza@becurio.com", "dinok@becurio.com", "whvkek@becurio.com"]

    search_target = date_str if date_str else "최근메일자동추출"
    print(f"\n--- [{venue_kor}] 한국 리포트 생성 파이프라인 시작 ({search_target}) ---")
    
    pred_df, extracted_date = fetch_prediction_table(venue_kor, date_str)
    
    if pred_df is not None and extracted_date is not None:
        print(f"[{venue_kor}] 대상 날짜 {extracted_date} 메일을 발견했습니다. DB 조회를 시작합니다.")
        date_hyphen = f"{extracted_date[:4]}-{extracted_date[4:6]}-{extracted_date[6:8]}"
        final_df = generate_hybrid_report(pred_df, venue_kor, date_hyphen)
        send_report_email(final_df, to_emails, venue_kor, extracted_date, cc_emails=cc_emails)
        return True
    else:
        print(f"⚠️ [{venue_kor}] 경주결과 Report 예측 메일을 찾을 수 없어 리포트 생성을 건너뜁니다.")
        return False

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="한국 결과 리포트 단독 실행")
    parser.add_argument("venue", help="경기장 한글명 (예: 서울, 부산, 제주)")
    parser.add_argument("date", nargs="?", help="날짜 (형식: YYYYMMDD, 생략 시 최근 메일에서 자동 추출)", default=None)
    
    args = parser.parse_args()
    run_kor_reporting_pipeline(args.venue, args.date)
