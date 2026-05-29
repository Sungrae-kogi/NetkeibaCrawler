import imaplib
import email
import json
import os
import smtplib
import pandas as pd
import pymysql
from email.header import decode_header
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from pathlib import Path

# 이 파일 기준으로 프로젝트 루트 경로 계산
# src/reporting/Reporting/ -> src/reporting/ -> src/ -> project_root/
_BASE_DIR = Path(__file__).resolve().parent
_PROJECT_ROOT = _BASE_DIR.parent.parent.parent
from io import StringIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

def get_credentials():
    config_path = _PROJECT_ROOT / "config" / "naverworks_config.json"
    if not config_path.exists():
        return None, None
    with open(config_path, "r", encoding="utf-8") as f:
        config = json.load(f)
        return config.get("email"), config.get("password")

def get_db_connection():
    import os
    env = os.environ.get("APP_ENV", "prod")
    config_name = "db_config_test.json" if env == "test" else "db_config.json"
    config_path = _PROJECT_ROOT / "config" / config_name
    with open(config_path, "r", encoding="utf-8") as f:
        config = json.load(f)
    return pymysql.connect(
        host=config['host'],
        port=config.get('port', 3306),
        user=config['user'],
        password=config['password'],
        database=config['database'],
        cursorclass=pymysql.cursors.DictCursor
    )

def decode_mime_words(s):
    if not s:
        return ""
    decoded_words = decode_header(s)
    result = []
    for word, charset in decoded_words:
        if isinstance(word, bytes):
            result.append(word.decode(charset or 'utf-8', errors='replace'))
        else:
            result.append(word)
    return "".join(result)

def fetch_prediction_table(venue_kor, date_str):
    user_email, user_password = get_credentials()
    imap_server = "imap.worksmobile.com"
    
    mail = imaplib.IMAP4_SSL(imap_server, 993)
    mail.login(user_email, user_password)
    mail.select("mafeel")
    
    status, messages = mail.search(None, 'FROM', '"mafeel@becurio.com"')
    mail_ids = messages[0].split()
    
    KOR_TO_JPN = {'교토': '京都', '도쿄': '東京', '나카야마': '中山', '한신': '阪神', '니가타': '新潟', '고쿠라': '小倉', '삿포로': '札幌', '하코다테': '函館', '후쿠시마': '福島', '중경': '中京'}
    venue_jp = KOR_TO_JPN.get(venue_kor, venue_kor)
    
    target_subject_part = f"[{venue_jp}-{date_str}]"
    
    df = None
    for i in reversed(mail_ids):
        res, msg_data = mail.fetch(i, "(RFC822)")
        for response_part in msg_data:
            if isinstance(response_part, tuple):
                msg = email.message_from_bytes(response_part[1])
                subject = decode_mime_words(msg["Subject"])
                if subject and target_subject_part in subject:
                    for part in msg.walk():
                        if part.get_content_maintype() == 'multipart': continue
                        if part.get_content_type() == "text/html" and part.get('Content-Disposition') is None:
                            try:
                                html_body = part.get_payload(decode=True).decode('utf-8', errors='replace')
                                tables = pd.read_html(StringIO(html_body), header=0)
                                for table in tables:
                                    # 예측 데이터 표는 보통 '순위', '출발번호', '한글마명' 등의 컬럼을 가짐
                                    if '순위' in table.columns or '한글마명' in table.columns:
                                        df = table
                                        break
                                if df is not None:
                                    break
                            except Exception:
                                pass
                    if df is not None:
                        mail.logout()
                        return df, venue_jp
    mail.logout()
    return None, venue_jp

def generate_hybrid_report(pred_df, venue_jp, date_hyphen):
    keep_cols = ['경주정보', '날짜', '경주번호', '순위', '출발번호', '한글마명']
    df = pred_df[keep_cols].copy()
    
    df.rename(columns={'순위': '[예측]순위', '출발번호': '[예측]출발번호', '한글마명': '[예측]한글마명'}, inplace=True)
    
    conn = get_db_connection()
    try:
        actual_data = []
        pred_odds_list = []
        pred_actual_ranks = []
        pred_popularity_list = []
        
        for _, row in df.iterrows():
            race_no = row['경주번호']
            pred_rank = row['[예측]순위']
            pred_horse_no = row['[예측]출발번호']
            
            with conn.cursor() as cursor:
                # 1. 예측마의 배당률, 실제 순위, 및 단승인기 가져오기
                race_suffix = f"%{str(race_no).zfill(2)}"
                sql_pred = "SELECT WIN_ODDS, RK, POPULARITY FROM tmp_races WHERE RCDATE=%s AND MEET=%s AND RCNO LIKE %s AND CHULNO=%s"
                cursor.execute(sql_pred, (date_hyphen, venue_jp, race_suffix, pred_horse_no))
                res_pred = cursor.fetchone()
                if res_pred:
                    pred_odds_list.append(res_pred['WIN_ODDS'])
                    pred_actual_rank = f"{res_pred['RK']}등" if res_pred['RK'] else '-'
                    pred_popularity_list.append(res_pred['POPULARITY'] if res_pred['POPULARITY'] else '-')
                else:
                    pred_odds_list.append('-')
                    pred_actual_rank = '-'
                    pred_popularity_list.append('-')
                pred_actual_ranks.append(pred_actual_rank)

            with conn.cursor() as cursor:
                # 2. 실제 순위 정보 대조 (1등, 2등, 3등 포디움 채우기 + 실제인기 추가)
                if pred_rank in ['1등', '2등', '3등']:
                    rank_num = pred_rank.replace('등', '')
                    sql_actual = "SELECT RK, CHULNO, HRNAME, WIN_ODDS, POPULARITY FROM tmp_races WHERE RCDATE=%s AND MEET=%s AND RCNO LIKE %s AND RK=%s"
                    cursor.execute(sql_actual, (date_hyphen, venue_jp, race_suffix, rank_num))
                    res_act = cursor.fetchone()
                    if res_act:
                        actual_data.append({
                            '[실제]순위': f"{res_act['RK']}등", 
                            '[실제]출발번호': res_act['CHULNO'], 
                            '[실제]한글마명': res_act['HRNAME'], 
                            '[실제]인기': res_act['POPULARITY'] if res_act['POPULARITY'] else '-',
                            '[실제]배당률': res_act['WIN_ODDS']
                        })
                    else:
                        actual_data.append({'[실제]순위': '-', '[실제]출발번호': '-', '[실제]한글마명': '-', '[실제]인기': '-', '[실제]배당률': '-'})
                else:
                    # 복병마 행은 포디움 칸 비워둠
                    actual_data.append({'[실제]순위': '-', '[실제]출발번호': '-', '[실제]한글마명': '-', '[실제]인기': '-', '[실제]배당률': '-'})
    finally:
        conn.close()
        
    df['[예측]인기'] = pred_popularity_list
    df['[예측]배당률'] = pred_odds_list
    df['[예측]실제순위'] = pred_actual_ranks
    actual_df = pd.DataFrame(actual_data)
    final_df = pd.concat([df.reset_index(drop=True), actual_df.reset_index(drop=True)], axis=1)
    return final_df

def create_excel_file(df, filepath):
    wb = Workbook()
    ws = wb.active
    ws.title = "결과보고서"
    
    headers_row1 = ["경주정보", "날짜", "경주번호", 
                    "예측 (Prediction)", "", "", "", "", "",
                    "실제 결과 (Actual)", "", "", "", ""]
    headers_row2 = ["", "", "", 
                    "순위", "출발번호", "한글마명", "인기", "배당률", "실제순위",
                    "순위", "출발번호", "한글마명", "인기", "배당률"]
    ws.append(headers_row1)
    ws.append(headers_row2)
    
    for _, row in df.iterrows():
        ws.append([
            row["경주정보"], row["날짜"], row["경주번호"],
            row["[예측]순위"], row["[예측]출발번호"], row["[예측]한글마명"], row["[예측]인기"], row["[예측]배당률"], row["[예측]실제순위"],
            row["[실제]순위"], row["[실제]출발번호"], row["[실제]한글마명"], row["[실제]인기"], row["[실제]배당률"]
        ])
        
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_header = PatternFill("solid", fgColor="EFEFEF")
    fill_hit = PatternFill("solid", fgColor="FFFF00") # 노란색 강조
    align_center = Alignment(horizontal="center", vertical="center")
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=14):
        for cell in row:
            cell.border = border
            cell.alignment = align_center
            if cell.row in [1, 2]:
                cell.fill = fill_header
                cell.font = Font(bold=True)
                
    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:B2")
    ws.merge_cells("C1:C2")
    ws.merge_cells("D1:I1") # 예측 영역 머지 (6개 열)
    ws.merge_cells("J1:N1") # 실제 결과 영역 머지 (5개 열)
    
    thick_bottom = Side(border_style="medium", color="000000")
                
    start_row = 3
    for i in range(3, ws.max_row + 2):
        if i > ws.max_row or ws.cell(row=i, column=3).value != ws.cell(row=i-1, column=3).value:
            end_row = i - 1
            
            # 예측마가 실제 1, 2, 3등 안에 들었으면 해당 행 노란색 강조
            for r in range(start_row, end_row + 1):
                actual_rank = str(ws.cell(row=r, column=9).value) # Column I: [예측]실제순위
                if actual_rank in ['1등', '2등', '3등']:
                    # 예측 컬럼들 D~I(4~9열) 노란색 칠하기
                    for col in range(4, 10):
                        ws.cell(row=r, column=col).fill = fill_hit

            # 경주 정보 셀 병합 (경주번호 기준)
            if end_row > start_row:
                ws.merge_cells(start_row=start_row, end_row=end_row, start_column=1, end_column=1)
                ws.merge_cells(start_row=start_row, end_row=end_row, start_column=2, end_column=2)
                ws.merge_cells(start_row=start_row, end_row=end_row, start_column=3, end_column=3)
                
            # 경주별 마지막 행(end_row)에 굵은 아래쪽 테두리 추가하여 시각적 구분 강화
            for col in range(1, 15):
                current_border = ws.cell(row=end_row, column=col).border
                ws.cell(row=end_row, column=col).border = Border(
                    left=current_border.left,
                    right=current_border.right,
                    top=current_border.top,
                    bottom=thick_bottom
                )
                
            start_row = i
            
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 10 # [예측]인기
    ws.column_dimensions['H'].width = 12 # [예측]배당률
    ws.column_dimensions['I'].width = 15 # [예측]실제순위
    ws.column_dimensions['J'].width = 12 # [실제]순위
    ws.column_dimensions['K'].width = 12 # [실제]출발번호
    ws.column_dimensions['L'].width = 20 # [실제]한글마명
    ws.column_dimensions['M'].width = 10 # [실제]인기
    ws.column_dimensions['N'].width = 12 # [실제]배당률
            
    wb.save(filepath)

def send_report_email(final_df, to_emails, venue_jp, date_str, cc_emails=None):
    user_email, user_password = get_credentials()
    smtp_server = "smtp.worksmobile.com"
    
    msg = MIMEMultipart()
    msg['From'] = user_email
    
    # 받는 사람 (To)
    if isinstance(to_emails, list):
        msg['To'] = ", ".join(to_emails)
    else:
        msg['To'] = to_emails
        
    # 참조 (Cc)
    if cc_emails:
        if isinstance(cc_emails, list):
            msg['Cc'] = ", ".join(cc_emails)
        else:
            msg['Cc'] = cc_emails
        
    # 제목: [京都-20260510] 예측과 실제 결과 비교 리포트
    msg['Subject'] = f"[{venue_jp}-{date_str}] 예측과 실제 결과 비교 리포트"
    
    # HTML 테이블 생성 (노란색 하이라이트 반영)
    html_lines = ['<table border="1" style="border-collapse: collapse; width: 100%; text-align: center;">', "  <thead>", "    <tr>"]
    for col in final_df.columns:
        html_lines.append(f'      <th style="background-color: #f2f2f2; padding: 8px;">{col}</th>')
    html_lines.append("    </tr>")
    html_lines.append("  </thead>")
    html_lines.append("  <tbody>")
    
    total_rows = len(final_df)
    for idx, (_, row) in enumerate(final_df.iterrows()):
        html_lines.append("    <tr>")
        actual_rank = str(row.get('[예측]실제순위', ''))
        is_hit = actual_rank in ['1등', '2등', '3등']
        
        # 다음 행의 경주번호가 다른지 판단 (마지막 행이거나 다음 경주번호와 다르면 경주 종료선 표시)
        is_last_of_race = False
        if idx == total_rows - 1:
            is_last_of_race = True
        else:
            next_race = final_df.iloc[idx + 1].get('경주번호')
            current_race = row.get('경주번호')
            if next_race != current_race:
                is_last_of_race = True
        
        for col in final_df.columns:
            val = row[col] if pd.notna(row[col]) else ''
            
            # 스타일 요소 정의
            styles = ["padding: 8px;"]
            if is_hit and col in ['[예측]순위', '[예측]출발번호', '[예측]한글마명', '[예측]인기', '[예측]배당률', '[예측]실제순위']:
                styles.append("background-color: #FFFF00;")
                
            if is_last_of_race:
                styles.append("border-bottom: 3px solid #000000;")
                
            style_str = f' style="{" ".join(styles)}"'
            html_lines.append(f"      <td{style_str}>{val}</td>")
        html_lines.append("    </tr>")
    html_lines.append("  </tbody>")
    html_lines.append("</table>")
    table_html = "\n".join(html_lines)
    
    date_hyphen = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
    html_content = f"""
    <html>
    <head>
        <style>
            table {{
                border-collapse: collapse;
                width: 100%;
                max-width: 1000px;
                font-family: 'Malgun Gothic', sans-serif;
                font-size: 12px;
                border: 1px solid #dddddd;
                margin-top: 15px;
                margin-bottom: 15px;
            }}
            th, td {{
                padding: 8px;
                text-align: center;
                border: 1px solid #dddddd;
            }}
            th {{
                background-color: #f2f2f2;
                font-weight: bold;
                color: #333333;
            }}
            tr:nth-child(even) {{
                background-color: #f9f9f9;
            }}
        </style>
    </head>
    <body>
        <p>{date_hyphen} {venue_jp} 경기 결과와 예측 비교 리포트를 첨부 파일로 송부합니다.</p>
        <br/>
        <h3>📊 [결과 비교 요약 표]</h3>
        {table_html}
        <br/>
        <p>감사합니다.</p>
    </body>
    </html>
    """
    msg.attach(MIMEText(html_content, 'html'))
    
    # 파일명: 京都_20260510_예측과_실제_결과_비교_리포트.xlsx
    excel_filename = f"{venue_jp}_{date_str}_예측과_실제_결과_비교_리포트.xlsx"
    create_excel_file(final_df, excel_filename)
    
    with open(excel_filename, "rb") as f:
        attach = MIMEApplication(f.read(), _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        attach.add_header('Content-Disposition', 'attachment', filename=excel_filename)
        msg.attach(attach)
    
    try:
        server = smtplib.SMTP_SSL(smtp_server, 465)
        server.login(user_email, user_password)
        server.send_message(msg)
        server.quit()
        print(f"[{venue_jp}-{date_str}] 이메일 발송 성공!")
    except Exception as e:
        print(f"이메일 발송 실패: {e}")

def run_reporting_pipeline(venue_jp, date_str, to_emails=None, cc_emails=None):
    """
    all.py 등 외부에서 호출하기 위한 메인 함수
    venue_jp: '京都', '東京' 등 (한자명)
    date_str: '20260510' 등
    to_emails: 받는 사람 이메일 (문자열 또는 리스트)
    cc_emails: 참조 이메일 (문자열 또는 리스트)
    """
    import os
    env = os.environ.get("APP_ENV", "prod")
    if env == "test":
        to_emails = ["whvkek@becurio.com"]
        cc_emails = None
        print("[TEST MODE] 이메일 수신자 격리 활성화 -> whvkek@becurio.com 수신 전용")
    else:
        if to_emails is None:
            # 받는 사람 (2명) : 오영섭 사장님, 일본 노승균 해외사업 본부장님
            to_emails = ["ysoh@becurio.com", "rhopritv65@gmail.com"]
        
        if cc_emails is None:
            # 참조 (3명)
            cc_emails = ["pizza@becurio.com", "dinok@becurio.com", "whvkek@becurio.com"]

    print(f"\n--- [{venue_jp}-{date_str}] 리포트 생성 파이프라인 시작 ---")
    pred_df, _ = fetch_prediction_table(venue_jp, date_str)
    
    if pred_df is not None:
        date_hyphen = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
        final_df = generate_hybrid_report(pred_df, venue_jp, date_hyphen)
        send_report_email(final_df, to_emails, venue_jp, date_str, cc_emails=cc_emails)
        return True
    else:
        print(f"⚠️ [{venue_jp}-{date_str}] 해당 경기의 예측 메일을 찾을 수 없어 리포트 생성을 건너뜁니다.")
        return False

if __name__ == "__main__":
    # 단독 실행 테스트용
    run_reporting_pipeline("京都", "20260510")
